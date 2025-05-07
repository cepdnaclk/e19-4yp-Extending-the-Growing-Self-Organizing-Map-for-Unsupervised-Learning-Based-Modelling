import os
import pandas as pd
from pandas import ExcelWriter
import numpy as np
from minisom import MiniSom
import gsom  # Assuming GSOM is a custom module
from openpyxl import load_workbook, Workbook
from scipy.spatial.distance import euclidean, cdist
import uuid
import zipfile

def zrehen_measure(weights, lattice_positions, neighbors, neurons):
    """
    Implements Zrehen Measure (ZM) as provided (Page 7).
    """
    N = len(weights)
    intruder_count = 0
    for i in range(N):
        for j in neighbors[i]:
            if j > i:
                w_i, w_j = weights[i], weights[j]
                dist_ij_sq = euclidean(w_i, w_j) ** 2
                for k in range(N):
                    if k != i and k != j:
                        w_k = weights[k]
                        dist_ik_sq = euclidean(w_i, w_k) ** 2
                        dist_jk_sq = euclidean(w_j, w_k) ** 2
                        if dist_ik_sq + dist_jk_sq <= dist_ij_sq:
                            intruder_count += 1
    zm = intruder_count / N
    return zm

def c_measure(weights, lattice_positions, neurons, p=2):
    """
    Implements C-Measure (CM) as provided (Page 8).
    """
    N = len(weights)
    cost = 0.0
    input_distances = cdist(weights, weights, metric='euclidean')
    map_distances = cdist(lattice_positions, lattice_positions, metric='euclidean') ** p
    for i in range(N):
        for j in range(N):
            if i != j:
                cost += input_distances[i, j] * map_distances[i, j]
    cm = cost / neurons
    return cm

def topographic_product(weights, lattice_positions):
    """
    Implements Topographic Product (TP) as provided (Page 8–9).
    """
    N = len(weights)
    total_log = 0.0
    input_distances = cdist(weights, weights, metric='euclidean')
    map_distances = cdist(lattice_positions, lattice_positions, metric='euclidean')
    for j in range(N):
        input_neighbors = np.argsort(input_distances[j])[1:]
        map_neighbors = np.argsort(map_distances[j])[1:]
        for k in range(1, N):
            product = 1.0
            for i in range(1, k+1):
                n_v = input_neighbors[i-1]
                n_a = map_neighbors[i-1]
                num = input_distances[j, n_v] * map_distances[j, n_a]
                denom = input_distances[j, n_a] * map_distances[j, n_v]
                if denom > 0:
                    product *= num / denom
            if product > 0:
                total_log += np.log(product ** (1 / (2 * k)))
    tp = total_log / (N**2 - N)
    return tp

def topographic_error(data, weights, lattice_positions, neighbors):
    """
    Implements Topographic Error (TE) as provided (Page 9).
    """
    K = len(data)
    error_count = 0
    distances = cdist(data, weights, metric='euclidean')
    for i in range(K):
        sorted_indices = np.argsort(distances[i])
        bmu1, bmu2 = sorted_indices[0], sorted_indices[1]
        if bmu2 not in neighbors[bmu1]:
            error_count += 1
    te = error_count / K
    return te

def get_neighbors(lattice_positions, grid_shape):
    """
    Computes neighbors for a 2D rectangular lattice (SOM), as provided (Page 4).
    """
    rows, cols = grid_shape
    N = len(lattice_positions)
    neighbors = [[] for _ in range(N)]
    for i in range(N):
        x, y = lattice_positions[i]
        for dx, dy in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
            nx, ny = x + dx, y + dy
            if 0 <= nx < rows and 0 <= ny < cols:
                neighbor_idx = np.where((lattice_positions[:, 0] == nx) & 
                                       (lattice_positions[:, 1] == ny))[0]
                if len(neighbor_idx) > 0:
                    neighbors[i].append(neighbor_idx[0])
    return neighbors

def get_neighbors_gsom(gsom_map):
    """
    Computes neighbors for GSOM based on grown connections (Page 2).
    """
    node_coords = gsom_map.node_coordinate[:gsom_map.node_count]
    N = gsom_map.node_count
    neighbors = [[] for _ in range(N)]
    for i in range(N):
        x_i, y_i = node_coords[i]
        for j in range(N):
            if i != j:
                x_j, y_j = node_coords[j]
                if (abs(x_i - x_j) == 1 and y_i == y_j) or (abs(y_i - y_j) == 1 and x_i == x_j):
                    neighbors[i].append(j)
    return neighbors

def load_dataset(file_path, is_3d=False):
    """
    Loads a dataset file (CSV) into a NumPy array.
    """
    df = pd.read_csv(file_path)
    cols = ['x', 'y', 'z'] if is_3d else ['x', 'y']
    data = df[cols[:2 if not is_3d else 3]].values
    # Normalize to [0,1]
    data = (data - data.min(axis=0)) / (data.max(axis=0) - data.min(axis=0) + 1e-10)
    return data

def is_valid_excel(file_path):
    """
    Check if the Excel file is a valid .xlsx file by verifying the presence of [Content_Types].xml.
    """
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            return '[Content_Types].xml' in z.namelist()
    except zipfile.BadZipFile:
        return False

def main(dataset_name, is_3d=False):
    """
    Tests topology preservation measures (ZM, CM, TP, TE) on a single dataset
    using SOM (minisom) and GSOM (gsom library), saving results and parameters to Excel.
    """
    # SOM and GSOM parameters
    grid_size = 100  # Size of the grid for SOM 
    grid_shape = (grid_size, grid_size)
    som_iterations = 60
    som_learning_rate = 0.01
    gsom_spread_factor = 0.5
    gsom_learning_rate = 0.01
    gsom_iterations = 60
    gsom_smoothing = 30
    gsom_max_radius = 4

    # Define parameters
    parameters = {
        'Run ID': str(uuid.uuid4()),  # Unique ID for each run
        'Grid Size': grid_size,
        'SOM Iterations': som_iterations,
        'SOM Learning Rate': som_learning_rate,
        'GSOM Spread Factor': gsom_spread_factor,
        'GSOM Learning Rate': gsom_learning_rate,
        'GSOM Iterations': gsom_iterations,
        'GSOM Smoothing': gsom_smoothing,
        'GSOM Max Radius': gsom_max_radius
    }

    # Create directories for datasets
    os.makedirs('data/shapes', exist_ok=True)
    os.makedirs('data/fcps', exist_ok=True)

    # Determine dataset type and path
    shapes_datasets = ['circle', 'triangle', 'rectangle', 'square', 'trapezoid', 'hexagon']
    file_path = f'data/shapes/{dataset_name}.csv' if dataset_name in shapes_datasets else f'data/fcps/{dataset_name}.csv'

    # Load data
    if not os.path.exists(file_path):
        print(f"Dataset {dataset_name} not found at {file_path}. Please download or provide.")
        return
    data = load_dataset(file_path, is_3d)

    # Train SOM
    som = MiniSom(grid_size, grid_size, data.shape[1], sigma=1.0, learning_rate=som_learning_rate)
    som.train_random(data, som_iterations)
    weights_som = som.get_weights().reshape(-1, data.shape[1])
    lattice_som = np.array([[i, j] for i in range(grid_size) for j in range(grid_size)])
    neighbors_som = get_neighbors(lattice_som, grid_shape)

    # Train GSOM
    gsom_map = gsom.GSOM(spred_factor=gsom_spread_factor, 
                         dimensions=data.shape[1], 
                         learning_rate=gsom_learning_rate, 
                         max_radius=gsom_max_radius)
    gsom_map.fit(data, gsom_iterations, gsom_smoothing)
    weights_gsom = gsom_map.node_list[:gsom_map.node_count]
    lattice_gsom = gsom_map.node_coordinate[:gsom_map.node_count]
    neighbors_gsom = get_neighbors_gsom(gsom_map)
    neurons = gsom_map.node_count

    # Compute topology preservation measures
    results = []
    zm_som = zrehen_measure(weights_som, lattice_som, neighbors_som, grid_size**2)
    cm_som = c_measure(weights_som, lattice_som, grid_size**2, p=2)
    tp_som = topographic_product(weights_som, lattice_som)
    te_som = topographic_error(data, weights_som, lattice_som, neighbors_som)

    zm_gsom = zrehen_measure(weights_gsom, lattice_gsom, neighbors_gsom, neurons)
    cm_gsom = c_measure(weights_gsom, lattice_gsom, neurons, p=2)
    tp_gsom = topographic_product(weights_gsom, lattice_gsom)
    te_gsom = topographic_error(data, weights_gsom, lattice_gsom, neighbors_gsom)

    # Append results to the list
    results.append({
        'Run ID': parameters['Run ID'],
        'Dataset': dataset_name,
        'Model': 'SOM',
        'Grid Size': grid_size,
        'ZM': zm_som,
        'CM': cm_som,
        'TP': tp_som,
        'TE': te_som
    })
    results.append({
        'Run ID': parameters['Run ID'],
        'Dataset': dataset_name,
        'Model': 'GSOM',
        'Neurons': neurons,
        'ZM': zm_gsom,
        'CM': cm_gsom,
        'TP': tp_gsom,
        'TE': te_gsom
    })

    # Print results
    print(f"\nDataset: {dataset_name}")
    print("SOM Results:")
    print(f"  ZM: {zm_som:.4f}, CM: {cm_som:.4f}, TP: {tp_som:.4f}, TE: {te_som:.4f}")
    print("GSOM Results:")
    print(f"  ZM: {zm_gsom:.4f}, CM: {cm_gsom:.4f}, TP: {tp_gsom:.4f}, TE: {te_gsom:.4f}")

    # Print parameters
    print("\nParameters:")
    for key, value in parameters.items():
        print(f"  {key}: {value}")
    

    # Combine parameters and results into a single DataFrame
    # parameters_df = pd.DataFrame([parameters])
    # results_df = pd.DataFrame(results)

    # # Merge parameters and results on Run ID
    # combined_df = results_df.merge(parameters_df, on='Run ID', how='left')

    # # Save to Excel
    # excel_file = 'topology_measures_results.xlsx'
    # sheet_name = dataset_name

    # try:
    #     # Check if file exists and is valid
    #     if os.path.exists(excel_file) and is_valid_excel(excel_file):
    #         book = load_workbook(excel_file)
    #         # If sheet exists, append data
    #         if sheet_name in book.sheetnames:
    #             existing_df = pd.read_excel(excel_file, sheet_name=sheet_name)
    #             # Ensure column consistency
    #             missing_cols = [col for col in combined_df.columns if col not in existing_df.columns]
    #             for col in missing_cols:
    #                 existing_df[col] = pd.NA
    #             # Concatenate existing data with new data
    #             final_df = pd.concat([existing_df, combined_df], ignore_index=True)
    #             # Remove the existing sheet
    #             del book[sheet_name]
    #         else:
    #             final_df = combined_df
    #     else:
    #         # If file is corrupted or doesn't exist, create a new workbook
    #         print(f"Excel file '{excel_file}' is corrupted or missing. Creating a new file.")
    #         book = Workbook()
    #         final_df = combined_df

    #     # Write to Excel
    #     with ExcelWriter(excel_file, engine='openpyxl') as writer:
    #         writer.book = book
    #         final_df.to_excel(writer, sheet_name=sheet_name, index=False)
    #         # Ensure at least one sheet exists
    #         if not book.sheetnames:
    #             book.create_sheet(sheet_name)

    #     print(f"\nParameters and results appended to '{excel_file}' in sheet '{sheet_name}'")
    # except Exception as e:
    #     print(f"Error saving to Excel: {e}")

if __name__ == "__main__":
    # Example: Test a single dataset
    main(dataset_name='Atom', is_3d=False)
    # main(dataset_name='hepta', is_3d=True)